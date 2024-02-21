# allocation_module.py
import pandas as pd
from PyQt5.QtWidgets import QMessageBox, QFileDialog
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class AllocationModule:
    def __init__(self, parent):
        self.parent = parent

    def read_excel_and_generate_output(self, file_path, students_per_class):
        df = pd.read_excel(file_path, header=None)

        l1 = []
        l2 = []

        for i in range(0, df.shape[1], 2):
            l1.extend(df.iloc[:, i].dropna().tolist())
            l2.extend(df.iloc[:, i + 1].dropna().tolist())

        len_diff = len(l1) - len(l2)

        if len_diff > 0:
            num_to_append = len_diff // 2
            l2.extend(l1[-num_to_append:])
            l1 = l1[:-num_to_append]
        elif len_diff < 0:
            num_to_append = abs(len_diff) // 2
            l1.extend(l2[-num_to_append:])
            l2 = l2[:-num_to_append]

        if len_diff % 2 == 1:
            last_element_l2 = l2.pop()

        l2.reverse()

        l3 = [item for pair in zip(l1, l2) for item in pair]

        if len_diff % 2 == 1:
            l3.append(last_element_l2)

        return l3, students_per_class

    def generate_excel_output(self, l3, students_per_class, rows, columns):
        data = {'Student_ID': l3}
        df_output = pd.DataFrame(data)

        students_total = len(l3)
        classes = -(-students_total // students_per_class)

        if rows * columns < students_per_class:
            QMessageBox.critical(self.parent, 'Error', 'The specified number of rows and columns is not sufficient to accommodate the students per class.')
            return

        remaining_slots = students_per_class * classes - students_total
        l3 += ['NONE'] * remaining_slots

        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        filename, _ = QFileDialog.getSaveFileName(self.parent, "Save Excel file", "AllocatedStudents.xlsx", "Excel Files (*.xlsx);;All Files (*)", options=options)

        if filename:
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"

            wb = Workbook()

            for i in range(classes):
                start_row = i * students_per_class
                end_row = min((i + 1) * students_per_class, students_total)
                class_data = df_output.iloc[start_row:end_row]

                if not class_data.empty:
                    none_values = pd.DataFrame({'Student_ID': ['NONE'] * (students_per_class - len(class_data))})
                    class_data = pd.concat([class_data, none_values], ignore_index=True)

                    reshaped_data = class_data.values.reshape((rows, columns), order='F')
                    reshaped_df = pd.DataFrame(reshaped_data)

                    reshaped_df.columns = [f'Room {j + 1}' for j in range(reshaped_df.shape[1])]

                    ws = wb.create_sheet(title=f'Class_{i + 1}')

                    for col in range(1, reshaped_df.shape[1] + 1):
                        column_letter = get_column_letter(col)
                        ws.column_dimensions[column_letter].width = 19

                    for r in dataframe_to_rows(reshaped_df, index=False, header=True):
                        ws.append(r)

                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
                    cell = ws.cell(row=1, column=1)
                    cell.value = f'Room {i + 1}'
                    cell.alignment = Alignment(horizontal='center')

            default_sheet = wb.active
            wb.remove(default_sheet)

            wb.save(filename)
            QMessageBox.information(self.parent, 'Success', f'Excel file "{filename}" has been created successfully.')
